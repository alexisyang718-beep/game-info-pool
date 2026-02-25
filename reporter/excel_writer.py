"""
Excel 输出模块
生成格式化的 Excel 文件，保存到 data/ 目录，由 GitHub Actions 提交到仓库
"""

import os
import pandas as pd
from pathlib import Path
from datetime import datetime

DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)


# 列名映射（英文字段 → 中文表头）
CHART_COLUMNS = {
    "rank":        "排名",
    "name":        "应用名称",
    "artist":      "开发商",
    "genre":       "品类",
    "region_name": "地区",
    "store":       "商店",
    "chart_name":  "榜单类型",
    "score":       "评分",
    "installs":    "安装量",
    "fetch_date":  "采集日期",
}

CHANGE_COLUMNS = {
    "name":        "应用名称",
    "artist":      "开发商",
    "region_name": "地区",
    "store":       "商店",
    "chart_name":  "榜单类型",
    "change_type": "异动类型",
    "rank_today":  "今日排名",
    "rank_yesterday": "昨日排名",
    "rank_delta":  "变化幅度",
}


def _store_label(store_val: str) -> str:
    return "Google Play" if store_val == "google_play" else "App Store"


def _apply_styles(writer, sheet_name: str, df: pd.DataFrame, header_color: str = "4472C4"):
    """为 worksheet 添加基础样式：冻结首行、自动列宽、表头颜色"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 自动列宽
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # 数据行边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def write_daily_excel(chart_data: list[dict], changes: list[dict], ai_analysis: str, date_str: str) -> str:
    """
    生成每日 Excel 文件
    包含 4 个 sheet：今日榜单、异动记录、AI分析、说明
    返回文件路径
    """
    filepath = DATA_DIR / f"榜单日报_{date_str}.xlsx"

    # ── 今日榜单 ──────────────────────────────────────────
    chart_rows = []
    for app in chart_data:
        chart_rows.append({
            "排名":     app.get("rank", ""),
            "应用名称": app.get("name", ""),
            "开发商":   app.get("artist", ""),
            "品类":     app.get("genre", ""),
            "地区":     app.get("region_name", app.get("region", "")),
            "商店":     _store_label(app.get("store", "")),
            "榜单类型": app.get("chart_name", ""),
            "评分":     app.get("score", ""),
            "安装量":   app.get("installs", ""),
            "采集日期": app.get("fetch_date", date_str),
        })
    df_charts = pd.DataFrame(chart_rows) if chart_rows else pd.DataFrame(columns=list(CHART_COLUMNS.values()))

    # ── 异动记录 ──────────────────────────────────────────
    change_rows = []
    for c in changes:
        change_rows.append({
            "应用名称": c.get("name", ""),
            "开发商":   c.get("artist", ""),
            "地区":     c.get("region_name", c.get("region", "")),
            "商店":     _store_label(c.get("store", "")),
            "榜单类型": c.get("chart_name", ""),
            "异动类型": c.get("change_type", ""),
            "今日排名": c.get("rank_today", ""),
            "昨日排名": c.get("rank_yesterday", ""),
            "变化幅度": c.get("rank_delta", ""),
        })
    df_changes = pd.DataFrame(change_rows) if change_rows else pd.DataFrame(columns=list(CHANGE_COLUMNS.values()))

    # ── AI 分析 ───────────────────────────────────────────
    df_analysis = pd.DataFrame([{"日期": date_str, "AI市场解读": ai_analysis}])

    # ── 写入 Excel ────────────────────────────────────────
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_charts.to_excel(writer, sheet_name="今日榜单", index=False)
        df_changes.to_excel(writer, sheet_name="今日异动", index=False)
        df_analysis.to_excel(writer, sheet_name="AI分析", index=False)

        # 写说明页
        df_info = pd.DataFrame([
            {"项目": "报告日期",    "内容": date_str},
            {"项目": "榜单条数",    "内容": len(chart_data)},
            {"项目": "异动数量",    "内容": len(changes)},
            {"项目": "覆盖地区",    "内容": "美国/英国/德国/法国/日本/韩国/印尼/泰国/新加坡/越南"},
            {"项目": "数据来源",    "内容": "App Store RSS API / Google Play"},
            {"项目": "更新时间",    "内容": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")},
        ])
        df_info.to_excel(writer, sheet_name="说明", index=False)

        _apply_styles(writer, "今日榜单", df_charts, "4472C4")
        _apply_styles(writer, "今日异动", df_changes, "C00000")
        _apply_styles(writer, "AI分析",  df_analysis, "70AD47")
        _apply_styles(writer, "说明",    df_info,     "7F7F7F")

    print(f"[Excel] 已生成 {filepath}")
    return str(filepath)


def write_weekly_excel(all_week_changes: list[dict], latest_charts: list[dict],
                       weekly_summary: str, date_str: str) -> str:
    """
    生成每周 Excel 文件
    包含：本周异动汇总、各地区当前榜单、AI周报
    """
    filepath = DATA_DIR / f"手游周报_{date_str}.xlsx"

    # 本周异动汇总
    change_rows = []
    for c in all_week_changes:
        change_rows.append({
            "应用名称": c.get("name", ""),
            "开发商":   c.get("artist", ""),
            "地区":     c.get("region_name", c.get("region", "")),
            "商店":     _store_label(c.get("store", "")),
            "榜单类型": c.get("chart_name", ""),
            "异动类型": c.get("change_type", ""),
            "今日排名": c.get("rank_today", ""),
            "昨日排名": c.get("rank_yesterday", ""),
            "变化幅度": c.get("rank_delta", ""),
        })
    df_changes = pd.DataFrame(change_rows) if change_rows else pd.DataFrame()

    # 各地区 Top 10
    top10_rows = [
        {
            "排名":     app.get("rank", ""),
            "应用名称": app.get("name", ""),
            "开发商":   app.get("artist", ""),
            "品类":     app.get("genre", ""),
            "地区":     app.get("region_name", ""),
            "商店":     _store_label(app.get("store", "")),
            "榜单类型": app.get("chart_name", ""),
        }
        for app in latest_charts if app.get("rank", 999) <= 10
    ]
    df_top10 = pd.DataFrame(top10_rows) if top10_rows else pd.DataFrame()

    df_summary = pd.DataFrame([{"周报日期": date_str, "AI周报内容": weekly_summary}])

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_top10.to_excel(writer, sheet_name="各地区Top10", index=False)
        df_changes.to_excel(writer, sheet_name="本周异动汇总", index=False)
        df_summary.to_excel(writer, sheet_name="AI周报", index=False)

        _apply_styles(writer, "各地区Top10",  df_top10,   "4472C4")
        _apply_styles(writer, "本周异动汇总", df_changes, "C00000")
        _apply_styles(writer, "AI周报",       df_summary, "70AD47")

    print(f"[Excel] 已生成 {filepath}")
    return str(filepath)
